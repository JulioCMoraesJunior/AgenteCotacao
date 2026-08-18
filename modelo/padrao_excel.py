import openpyxl

planilha = openpyxl.load_workbook("C:/Users/Julio Moraes/Desktop/agente_automacao/teste.xlsx")

aba = planilha['Plan1']

produtos = []
for linha in range(2, aba.max_row +1):
    produto = aba.cell(linha, 1).value
    produtos.append(produto)

PRODUTOS_OFICIAIS = "\n".join(produtos)
