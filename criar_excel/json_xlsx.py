import openpyxl
import json

def excel():
    with open('json/resposta_llm.json', 'r', encoding = 'utf-8') as arquivo:
        resposta_llm = json.load(arquivo)
    planilha = openpyxl.load_workbook("teste.xlsx")
    aba = planilha['Plan1']

    coluna = 2
    for pedido in resposta_llm:
        usuario = pedido['usuario']
        aba.cell(1, coluna).value = usuario
        for produto_procurado, quantidade in pedido['produtos'].items():
            print(produto_procurado)
            print(quantidade)
            for linha in range(1, aba.max_row+1):
                    produto = aba.cell(linha, 1).value
                    if produto == produto_procurado:
                        aba.cell(linha, coluna).value = quantidade
        coluna += 1
    planilha.save("teste.xlsx")
